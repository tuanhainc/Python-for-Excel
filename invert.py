import openpyxl


backup = input('Press Enter to use sample file  ')
if len(backup)>1:
    file = input('File name: ')
else:
    file = 'example.xlsx'

try:
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
except:
    print('No file found')
    exit()

wbfinal = openpyxl.Workbook()
sheetfinal = wbfinal.active

sheetdata = {}
for rowNum in range(1,sheet.max_row+1):
    for colNum in range(1,sheet.max_column+1):
        value = sheet.cell(row=rowNum,column=colNum).value
        sheetdata.setdefault(rowNum,{})
        sheetdata[rowNum].setdefault(colNum,value)

# transpose
for rowNum in range(1,sheet.max_row+1):
    for colNum in range(1,sheet.max_column+1):
        sheetfinal.cell(row=colNum,column=rowNum).value = sheetdata[rowNum][colNum]


wbfinal.save('inverted.xlsx')


# python invert.py
