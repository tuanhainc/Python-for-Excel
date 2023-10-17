import openpyxl


backup = input('Press Enter to use sample file  ')
if len(backup)>1:
    N = int(input('At row number: '))-1
    M = int(input('How many blank row? '))
    file = input('File name: ')
else:
    N = 5
    M = 4
    file = 'example.xlsx'

wb = openpyxl.load_workbook(file)
sheet = wb.active

wbfinal = openpyxl.Workbook()
sheetfinal = wbfinal.active

for row in range(1,N+1):
    for cell in range(1, sheet.max_column+1):
        sheetfinal.cell(row=row, column=cell).value = sheet.cell(row=row, column=cell).value

for row in range(N+M,sheet.max_row+M+1):
    for cell in range(1, sheet.max_column+1):
        sheetfinal.cell(row=row+1, column=cell).value = sheet.cell(row=row-M+1, column=cell).value

wbfinal.save('file_blankrow.xlsx')


# python blankrowinserter.py
