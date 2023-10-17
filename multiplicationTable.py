import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb['Sheet']

N = int(input("N = "))

# write row colomn headers
bold = Font(bold=True)
for y in range (2, N + 2):
    sheet.cell(row=1, column=y).value = y - 1
    sheet.cell(row=1, column=y).font = bold
    sheet.cell(row=y, column=1).value = y - 1
    sheet.cell(row=y, column=1).font = bold
# write multiplicationTable
for row in range(2, N + 2):
    for col in range(2, N + 2):
        sheet.cell(row=row, column=col).value = (row-1)*(col-1)

wb.save('multiplicationTable.xlsx')



# python multiplicationTable.py
